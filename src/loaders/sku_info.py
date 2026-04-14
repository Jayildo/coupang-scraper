"""sku_info 마스터 업로드 (xlsx → UPSERT + content_hash 변경 감지).

CSV/xlsx 17컬럼:
  PK: sku_id
  Promoted: sku_name(상품명), status(발주가능상태)
  Rest: raw_data jsonb (모든 한국어 컬럼 보존)

알고리즘:
  1. xlsx 파싱 → 모든 컬럼을 dict
  2. 기존 active sku_id → content_hash 맵 페치
  3. 신규/변경/동일 분류 + sku_info upsert
  4. 사라진 sku → is_active=false (소프트 삭제)
"""
import logging
from datetime import datetime
from pathlib import Path

from loaders.base import (
    chunked, finish_run, get_supabase_client, latest_download,
    now_iso, stable_hash, start_run,
)

log = logging.getLogger(__name__)

# Excel 한국어 컬럼 → promoted DB 컬럼
PROMOTED = {
    "SKU ID": "sku_id",
    "상품명": "sku_name",
    "발주가능상태": "status",
    "바코드": "barcode_text",   # 단, schema에 barcode 없음 → raw_data로만 보존
}

# content_hash 계산 시 제외할 메타 컬럼 (없음 — 모든 비즈니스 컬럼 포함)
BATCH_SIZE = 500


def _xlsx_to_rows(path: Path):
    """xlsx 파싱 → row dict 리스트."""
    from openpyxl import load_workbook
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(2, ws.max_row + 1):
        raw = {}
        for c, h in enumerate(headers, start=1):
            if h is None:
                continue
            v = ws.cell(row=r, column=c).value
            # 모든 값을 string화 해서 raw_data에 저장 (jsonb numeric/string mixed 회피)
            raw[h] = "" if v is None else str(v)

        sku_id = raw.get("SKU ID", "").strip()
        if not sku_id:
            continue

        row = {
            "sku_id": sku_id,
            "sku_name": raw.get("상품명") or None,
            "status": raw.get("발주가능상태") or None,
            "raw_data": raw,
            "content_hash": stable_hash(raw),
        }
        rows.append(row)
    return rows


def _fetch_existing_active(client) -> dict:
    """sku_id → content_hash 맵."""
    existing = {}
    page_size = 1000
    offset = 0
    while True:
        resp = client.table("sku_info") \
            .select("sku_id,content_hash") \
            .eq("is_active", True) \
            .range(offset, offset + page_size - 1) \
            .execute()
        data = resp.data or []
        for r in data:
            existing[r["sku_id"]] = r["content_hash"]
        if len(data) < page_size:
            break
        offset += page_size
    return existing


def load(file_path: Path | None = None) -> dict:
    if file_path is None:
        file_path = latest_download("sku_info_*.xlsx")
    if not file_path or not Path(file_path).exists():
        raise FileNotFoundError(f"sku_info xlsx 파일 없음: {file_path}")

    file_path = Path(file_path)
    log.info(f"[SKU_INFO_LOAD] 파일: {file_path}")

    client = get_supabase_client()
    run_id = start_run(client, "sku_info", file_path)

    try:
        new_rows = _xlsx_to_rows(file_path)
        log.info(f"[SKU_INFO_LOAD] 파싱 완료: {len(new_rows)} rows")

        existing = _fetch_existing_active(client)
        log.info(f"[SKU_INFO_LOAD] 기존 active: {len(existing)} rows")

        new_sku_ids = {r["sku_id"] for r in new_rows}

        n_new = n_changed = n_same = 0
        now = now_iso()
        new_payloads = []
        changed_payloads = []
        unchanged_ids = []
        for row in new_rows:
            sku_id = row["sku_id"]
            old_hash = existing.get(sku_id)
            if old_hash is not None and old_hash == row["content_hash"]:
                unchanged_ids.append(sku_id)
                n_same += 1
                continue
            base = {
                **row,
                "scrape_run_id": run_id,
                "scraped_at": now,
                "last_seen_at": now,
                "is_active": True,
                "last_changed_at": now,
            }
            if old_hash is None:
                base["first_seen_at"] = now
                new_payloads.append(base)
                n_new += 1
            else:
                changed_payloads.append(base)
                n_changed += 1

        for label, payloads in (("new", new_payloads), ("changed", changed_payloads)):
            upserted = 0
            for batch in chunked(payloads, BATCH_SIZE):
                client.table("sku_info").upsert(batch, on_conflict="sku_id").execute()
                upserted += len(batch)
                log.info(f"[SKU_INFO_LOAD] {label} upserted {upserted}/{len(payloads)}")

        # unchanged: last_seen_at만 touch
        for chunk in chunked(unchanged_ids, 500):
            client.table("sku_info") \
                .update({"last_seen_at": now, "scrape_run_id": run_id}) \
                .in_("sku_id", chunk) \
                .execute()

        # 사라진 sku → is_active=false
        removed = list(set(existing.keys()) - new_sku_ids)
        n_removed = len(removed)
        if removed:
            log.info(f"[SKU_INFO_LOAD] 사라진 SKU {n_removed}건 → is_active=false")
            # in_() 필터로 일괄 update (단일 PK)
            for chunk in chunked(removed, 200):
                client.table("sku_info") \
                    .update({"is_active": False, "last_changed_at": now}) \
                    .in_("sku_id", chunk) \
                    .execute()

        finish_run(
            client, run_id,
            status="success",
            rows_processed=len(new_rows),
            rows_inserted=n_new,
            rows_updated=n_changed,
            rows_unchanged=n_same,
            rows_removed=n_removed,
            metadata={"file": str(file_path)},
        )
        return {
            "run_id": run_id,
            "rows": len(new_rows),
            "new": n_new, "changed": n_changed, "unchanged": n_same, "removed": n_removed,
        }

    except Exception as e:
        log.error(f"[SKU_INFO_LOAD] 실패: {e}", exc_info=True)
        finish_run(client, run_id, status="failed", error_message=str(e))
        raise
